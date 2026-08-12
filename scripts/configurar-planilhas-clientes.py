# -*- coding: utf-8 -*-
"""Aplica a configuracao da aba GERAL (planilha modelo) nas planilhas dos 18 clientes.

Nao apaga informacao: insere a coluna ORDEM, padroniza as formulas de PRAZO,
instala as cores e a lista de status, e reordena as linhas com edital pelo prazo.

Uso:  python configurar_clientes.py            (simulacao, nao grava)
      python configurar_clientes.py --aplicar  (grava, com backup antes)
"""
import sys, os, shutil, datetime, re
from copy import copy, deepcopy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

sys.stdout.reconfigure(encoding="utf-8")

APLICAR = "--aplicar" in sys.argv
HOJE = datetime.datetime(2026, 8, 11)

DESKTOP = r"C:\Users\rosep\Desktop\_82 - Rosepaula Aparecida Andrade Rodrigues"
MODELO = os.path.join(DESKTOP, r"04 - Controle de Submissão_\01 - Mineração de Editais\01 - Planejamento de Submissões\1 - Controle de Submissão_.xlsx")
CLIENTES = os.path.join(DESKTOP, "06 - Clientes")
BACKUP = os.path.join(CLIENTES, "_Backup planilhas 11-08-2026")

ULTIMA_LINHA = 300          # ate onde vao as formulas, igual ao modelo
ULTIMA_LINHA_DV = 1000      # ate onde vai a lista suspensa de status

STATUS_EXTRA = {
    "Selecionado": ("FF4DB6AC", "FF000000"),
    "Recurso":     ("FFEF6C00", "FFFFFFFF"),
}

# ------------------------------------------------------------------ modelo
wbm = openpyxl.load_workbook(MODELO)
wsm = wbm["GERAL"]

STATUS_OFICIAIS = [str(c.value) for (c,) in wsm.parent["Status"].iter_rows(min_row=2, max_row=26, max_col=1) if c.value]
LISTA_STATUS = STATUS_OFICIAIS + list(STATUS_EXTRA)

regras_status, regras_prazo = [], []
for cf in wsm.conditional_formatting:
    ref = str(cf.sqref)
    for r in cf.rules:
        if ref.startswith("L3"):
            regras_status.append(deepcopy(r))
        elif ref.startswith("H1 H3"):
            regras_prazo.append(deepcopy(r))

for texto, (fundo, fonte) in STATUS_EXTRA.items():
    regras_status.append(Rule(
        type="cellIs", operator="equal", formula=[f'"{texto}"'],
        dxf=DifferentialStyle(
            fill=PatternFill(start_color=fundo, end_color=fundo, fill_type="solid"),
            font=Font(color=fonte),
        ),
    ))

print(f"Modelo lido: {len(STATUS_OFICIAIS)} status oficiais + {len(STATUS_EXTRA)} acrescentados = {len(LISTA_STATUS)}")
print(f"Regras de cor do modelo: {len(regras_prazo)} de prazo, {len(regras_status)} de status")
print(f"Modo: {'APLICAR (grava os arquivos)' if APLICAR else 'SIMULACAO (nao grava nada)'}\n")

# ------------------------------------------------------------------ auxiliares
def copiar_celula(ws, lin_o, col_o, lin_d, col_d):
    o, d = ws.cell(lin_o, col_o), ws.cell(lin_d, col_d)
    d.value = o.value
    d._style = copy(o._style)
    d.number_format = o.number_format
    d.hyperlink = copy(o.hyperlink) if o.hyperlink else None

def limpar_celula(ws, lin, col):
    c = ws.cell(lin, col)
    c.value = None
    c.hyperlink = None

def mapa_cabecalho(ws, lin):
    m = {}
    for c in range(1, ws.max_column + 2):
        v = str(ws.cell(lin, c).value or "").strip().upper()
        if v:
            m.setdefault(v, []).append(c)
    return m

def chave_ordem(f):
    """Ordena: com data pelo prazo crescente, depois continuos, depois sem data."""
    if isinstance(f, datetime.datetime):
        return (0, (f - HOJE).days)
    if isinstance(f, str) and "ontinu" in f.lower():
        return (1, 0)
    return (2, 0)

# ------------------------------------------------------------------ backup
if APLICAR:
    os.makedirs(BACKUP, exist_ok=True)

relatorio = []
for pasta in sorted(os.listdir(CLIENTES)):
    cam = os.path.join(CLIENTES, pasta)
    if not os.path.isdir(cam) or pasta.startswith("_"):
        continue
    arqs = [f for f in os.listdir(cam) if "ontrole" in f and f.lower().endswith((".xlsx", ".xlsm"))]
    if not arqs:
        continue
    arq = arqs[0]
    caminho = os.path.join(cam, arq)
    avisos = []

    wb = openpyxl.load_workbook(caminho)
    nome_aba = next((n for n in wb.sheetnames if n.strip().lower().startswith("submiss")), None)
    renomeada = False
    if nome_aba is None:
        nome_aba = wb.sheetnames[0]
        renomeada = True
    ws = wb[nome_aba]

    # ---- 1. normaliza cabecalho na linha 2 e o titulo na linha 1
    lin_cab = 1 if str(ws.cell(1, 2).value or "").strip().upper() == "NOME" else 2
    if lin_cab == 1:
        ws.insert_rows(1)
        lin_cab = 2
        # estilo vindo de uma celula do proprio arquivo: indice de estilo nao atravessa arquivos
        ws.cell(1, 1).value = "CONTROLE DE SUBMISSÃO"
        ws.cell(1, 1)._style = copy(ws.cell(lin_cab, 2)._style)
        avisos.append("cabecalho movido para a linha 2 e titulo criado na linha 1")
    if renomeada:
        ws.title = "SUBMISSÕES"
        avisos.append(f"aba '{nome_aba}' renomeada para 'SUBMISSÕES'")

    cab = mapa_cabecalho(ws, lin_cab)
    col_nome = cab["NOME"][0]
    col_data = cab["DATA LIMITE"][0]
    col_prazo = cab["PRAZO"][0]
    col_cliente = cab["CLIENTE"][0]
    col_prox = cab["PRÓXIMA ETAPA"][0]
    tem_ordem = "ORDEM" in cab

    # desfaz a mesclagem do titulo, senao as celulas da linha 1 sao somente leitura
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    # ---- 2. insere a coluna ORDEM logo depois de PRAZO
    if not tem_ordem:
        col_ordem = col_prazo + 1
        ult = ws.max_column
        for c in range(ult, col_ordem - 1, -1):
            for r in range(1, ULTIMA_LINHA_DV + 1):
                copiar_celula(ws, r, c, r, c + 1)
        for r in range(1, ULTIMA_LINHA_DV + 1):
            limpar_celula(ws, r, col_ordem)
        ws.cell(lin_cab, col_ordem).value = "ORDEM"
        ws.cell(lin_cab, col_ordem)._style = copy(ws.cell(lin_cab, col_prazo)._style)
        ws.column_dimensions[get_column_letter(col_ordem)].width = 12
        ws.column_dimensions[get_column_letter(col_ordem)].hidden = True
        avisos.append(f"coluna ORDEM criada em {get_column_letter(col_ordem)} (colunas seguintes deslocadas)")
        cab = mapa_cabecalho(ws, lin_cab)
    else:
        col_ordem = cab["ORDEM"][0]
        ws.column_dimensions[get_column_letter(col_ordem)].hidden = True

    col_nome = cab["NOME"][0]
    col_data = cab["DATA LIMITE"][0]
    col_prazo = cab["PRAZO"][0]
    col_prox = cab["PRÓXIMA ETAPA"][0]
    col_status = cab["STATUS"][0]
    col_prazo2 = [c for c in cab["PRAZO"] if c != col_prazo]
    col_prazo2 = col_prazo2[0] if col_prazo2 else None
    ult_col = ws.max_column

    # ---- 3. reordena as linhas que tem edital, nas mesmas posicoes
    linhas = [r for r in range(lin_cab + 1, ULTIMA_LINHA + 1) if ws.cell(r, col_nome).value not in (None, "")]
    registros = []
    for r in linhas:
        cel = {}
        for c in range(1, ult_col + 1):
            o = ws.cell(r, c)
            cel[c] = (o.value, copy(o._style), o.number_format, copy(o.hyperlink) if o.hyperlink else None)
        registros.append({
            "cel": cel,
            "f": ws.cell(r, col_data).value,
            "nome": str(ws.cell(r, col_nome).value)[:40],
            "altura": ws.row_dimensions[r].height,
        })
    ordenados = sorted(registros, key=lambda x: chave_ordem(x["f"]))
    mudou = [i for i, (a, b) in enumerate(zip(registros, ordenados)) if a is not b]
    for destino, reg in zip(linhas, ordenados):
        for c in range(1, ult_col + 1):
            v, est, fmt, hl = reg["cel"][c]
            d = ws.cell(destino, c)
            d.value, d._style, d.number_format, d.hyperlink = v, est, fmt, hl
        if reg["altura"]:
            ws.row_dimensions[destino].height = reg["altura"]

    sem_data = [r["nome"] for r in registros if r["f"] in (None, "")]

    # ---- 4. formulas de PRAZO, ORDEM e prazo da proxima etapa
    L_data, L_prazo, L_prox = get_column_letter(col_data), get_column_letter(col_prazo), get_column_letter(col_prox)
    for r in range(lin_cab + 1, ULTIMA_LINHA + 1):
        ws.cell(r, col_prazo).value = (
            f'=IF({L_data}{r}="","",'
            f'IF(OR({L_data}{r}="Continuo",{L_data}{r}="Contínuo",{L_data}{r}="CONTINUO",'
            f'{L_data}{r}="CONTÍNUO",{L_data}{r}="Contìnuo"),"Contínuo",{L_data}{r}-TODAY()))'
        )
        ws.cell(r, col_ordem).value = (
            f'=IF({L_prazo}{r}="","",'
            f'IF(ISTEXT({L_prazo}{r}),9999,IF({L_prazo}{r}=0,9998,'
            f'IF({L_prazo}{r}<0,ABS({L_prazo}{r}),1000+{L_prazo}{r}))))'
        )
        if col_prazo2:
            ws.cell(r, col_prazo2).value = f'=IF({L_prox}{r}="","",{L_prox}{r}-TODAY())'

    # ---- 5. aba Status, intervalo nomeado e lista suspensa
    if "Status" in wb.sheetnames:
        del wb["Status"]
    wss = wb.create_sheet("Status")
    wss.cell(1, 1).value = "STATUS"
    for i, s in enumerate(LISTA_STATUS, start=2):
        wss.cell(i, 1).value = s
    wss.column_dimensions["A"].width = 32
    if "Lista_Status" in wb.defined_names:
        del wb.defined_names["Lista_Status"]
    wb.defined_names.add(DefinedName("Lista_Status", attr_text=f"Status!$A$2:$A${len(LISTA_STATUS)+1}"))

    ws.data_validations.dataValidation = []
    dv = DataValidation(type="list", formula1="Lista_Status", allow_blank=True)
    ws.add_data_validation(dv)
    L_status = get_column_letter(col_status)
    dv.add(f"{L_status}{lin_cab+1}:{L_status}{ULTIMA_LINHA_DV}")

    # ---- 6. cores: remove as mortas e as antigas de prazo/status, instala as do modelo
    def morta(regra):
        for f in (getattr(regra, "formula", None) or []):
            t = str(f)
            if "#REF!" in t or t.strip().startswith('" ='):
                return True
        return False

    novo = ConditionalFormattingList()
    removidas = 0
    for cf in ws.conditional_formatting:
        ref = str(cf.sqref)
        toca_prazo = re.search(rf"\b{L_prazo}\d", ref) or re.search(rf"\b{L_status}\d", ref)
        for regra in cf.rules:
            if morta(regra) or (toca_prazo and regra.type != "notContainsBlanks"):
                removidas += 1
                continue
            novo.add(ref, regra)
    ws.conditional_formatting = novo

    faixa_prazo = f"{L_prazo}{lin_cab+1}:{L_prazo}{ULTIMA_LINHA_DV}"
    faixa_status = f"{L_status}{lin_cab+1}:{L_status}{ULTIMA_LINHA_DV}"
    for r in regras_prazo:
        ws.conditional_formatting.add(faixa_prazo, deepcopy(r))
    for r in regras_status:
        ws.conditional_formatting.add(faixa_status, deepcopy(r))

    # ---- 7. titulo mesclado igual ao modelo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(ult_col, 16))

    if APLICAR:
        shutil.copy2(caminho, os.path.join(BACKUP, f"{pasta} - {arq}"))
        wb.save(caminho)

    relatorio.append({
        "cliente": pasta, "arquivo": arq, "aba": ws.title,
        "editais": len(linhas), "reordenadas": len(mudou),
        "ordem": get_column_letter(col_ordem), "status": L_status,
        "prazo2": get_column_letter(col_prazo2) if col_prazo2 else "-",
        "cf_removidas": removidas, "sem_data": sem_data, "avisos": avisos,
    })

# ------------------------------------------------------------------ relatorio
print(f"{'CLIENTE':<36} {'EDITAIS':>7} {'REORD':>6} {'ORDEM':>6} {'STATUS':>7} {'PRAZO2':>7} {'CF-':>4}")
print("-" * 82)
for r in relatorio:
    print(f"{r['cliente'][:36]:<36} {r['editais']:>7} {r['reordenadas']:>6} {r['ordem']:>6} {r['status']:>7} {r['prazo2']:>7} {r['cf_removidas']:>4}")
print("-" * 82)
print(f"{len(relatorio)} planilhas | {sum(r['editais'] for r in relatorio)} editais | {sum(r['reordenadas'] for r in relatorio)} linhas reposicionadas")

print("\nAVISOS POR CLIENTE:")
for r in relatorio:
    if r["avisos"] or r["sem_data"]:
        print(f"  {r['cliente']}")
        for a in r["avisos"]:
            print(f"     - {a}")
        for s in r["sem_data"]:
            print(f"     - sem DATA LIMITE, foi para o fim do bloco: {s}")
