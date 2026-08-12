# -*- coding: utf-8 -*-
"""Nova regra de ORDEM, decidida em 11/08/2026:
   abertos primeiro (prazo crescente), depois os continuos, e os vencidos no rodape
   (do vencimento mais recente para o mais antigo). Sem data limite fica por ultimo.

   Faixas da coluna ORDEM:  0 a 4999 = aberto (dias que faltam) | 5000 = continuo
                            6000 + dias vencidos = vencido      | vazio = sem data

Uso: python nova_ordem.py [--aplicar]
"""
import sys, os, datetime, openpyxl
from copy import copy
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
APLICAR = "--aplicar" in sys.argv
HOJE = datetime.datetime(2026, 8, 11)

DESKTOP = r"C:\Users\rosep\Desktop\_82 - Rosepaula Aparecida Andrade Rodrigues"
MODELO = os.path.join(DESKTOP, r"04 - Controle de Submissão_\01 - Mineração de Editais\01 - Planejamento de Submissões\1 - Controle de Submissão_.xlsx")
CLIENTES = os.path.join(DESKTOP, "06 - Clientes")
ULTIMA = 300

def formula_ordem(L_prazo, r):
    return (f'=IF({L_prazo}{r}="","",'
            f'IF(ISTEXT({L_prazo}{r}),5000,'
            f'IF({L_prazo}{r}>=0,{L_prazo}{r},6000+ABS({L_prazo}{r}))))')

def sem_acento(s):
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn").lower()

def eh_continuo(f):
    # precisa ignorar acento: 'Contínuo', 'Continuo', 'CONTÍNUO' e 'Contìnuo' aparecem todos na base
    return isinstance(f, str) and "continu" in sem_acento(f)

def chave(f):
    if isinstance(f, datetime.datetime):
        d = (f - HOJE).days
        return (0, d) if d >= 0 else (2, abs(d))
    if eh_continuo(f):
        return (1, 0)
    return (3, 0)

def colunas(ws, cab):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(cab, c).value or "").strip().upper()
        if v:
            m.setdefault(v, []).append(c)
    return m

def processar(caminho, aba, reordenar):
    wb = openpyxl.load_workbook(caminho)
    ws = wb[aba]
    cab = 2
    m = colunas(ws, cab)
    col_ordem = m["ORDEM"][0]
    col_prazo = m["PRAZO"][0]
    col_nome = m["NOME"][0]
    col_data = m["DATA LIMITE"][0]
    L_prazo = get_column_letter(col_prazo)

    for r in range(cab + 1, ULTIMA + 1):
        ws.cell(r, col_ordem).value = formula_ordem(L_prazo, r)

    trocadas = 0
    if reordenar:
        linhas = [r for r in range(cab + 1, ULTIMA + 1) if ws.cell(r, col_nome).value not in (None, "")]
        ult_col = ws.max_column
        regs = []
        for r in linhas:
            regs.append({
                "cel": {c: (ws.cell(r, c).value, copy(ws.cell(r, c)._style), ws.cell(r, c).number_format,
                            copy(ws.cell(r, c).hyperlink) if ws.cell(r, c).hyperlink else None)
                        for c in range(1, ult_col + 1)},
                "f": ws.cell(r, col_data).value,
                "altura": ws.row_dimensions[r].height,
            })
        ordenados = sorted(regs, key=lambda x: chave(x["f"]))
        trocadas = sum(1 for a, b in zip(regs, ordenados) if a is not b)
        for destino, reg in zip(linhas, ordenados):
            for c in range(1, ult_col + 1):
                v, est, fmt, hl = reg["cel"][c]
                d = ws.cell(destino, c)
                d.value, d._style, d.number_format, d.hyperlink = v, est, fmt, hl
            if reg["altura"]:
                ws.row_dimensions[destino].height = reg["altura"]
        # as colunas de calculo voltam para o formato Geral
        for c in m.get("PRAZO", []) + [col_ordem]:
            for r in range(cab + 1, ULTIMA + 1):
                ws.cell(r, c).number_format = "General"
        # e recebem a formula de novo, ja que o valor viajou junto com o registro
        L_data = get_column_letter(col_data)
        col_prox = m["PRÓXIMA ETAPA"][0]
        L_prox = get_column_letter(col_prox)
        col_prazo2 = [c for c in m["PRAZO"] if c != col_prazo]
        for r in range(cab + 1, ULTIMA + 1):
            ws.cell(r, col_prazo).value = (
                f'=IF({L_data}{r}="","",'
                f'IF(OR({L_data}{r}="Continuo",{L_data}{r}="Contínuo",{L_data}{r}="CONTINUO",'
                f'{L_data}{r}="CONTÍNUO",{L_data}{r}="Contìnuo"),"Contínuo",{L_data}{r}-TODAY()))')
            ws.cell(r, col_ordem).value = formula_ordem(L_prazo, r)
            if col_prazo2:
                ws.cell(r, col_prazo2[0]).value = f'=IF({L_prox}{r}="","",{L_prox}{r}-TODAY())'
    if APLICAR:
        wb.save(caminho)
    return len(linhas) if reordenar else 0, trocadas

n, t = processar(MODELO, "GERAL", reordenar=False)
print(f"{'MODELO (aba GERAL)':<36} formula ORDEM atualizada, ordem fisica preservada (blocos manuais)")

tot_e = tot_t = 0
for pasta in sorted(os.listdir(CLIENTES)):
    cam = os.path.join(CLIENTES, pasta)
    if not os.path.isdir(cam) or pasta.startswith("_"):
        continue
    arqs = [f for f in os.listdir(cam) if "ontrole" in f and f.lower().endswith(".xlsx")]
    if not arqs:
        continue
    wbtmp = openpyxl.load_workbook(os.path.join(cam, arqs[0]))
    aba = next(x for x in wbtmp.sheetnames if x.strip().lower().startswith("submiss"))
    e, t = processar(os.path.join(cam, arqs[0]), aba, reordenar=True)
    tot_e += e; tot_t += t
    print(f"{pasta[:36]:<36} editais={e:<3} reposicionados={t}")

print(f"\n{tot_e} editais, {tot_t} reposicionados. {'APLICADO' if APLICAR else 'SIMULACAO'}")
